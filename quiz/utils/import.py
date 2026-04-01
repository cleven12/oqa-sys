import openpyxl
from ..models import Question, QuestionGroup


def import_from_excel(quiz, excel_file):
    """
    Import questions from Excel file
    Returns: (success_count, error_list)
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    success_count = 0
    errors = []
    
    headers = [cell.value for cell in ws[1]]
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # Skip empty rows
                continue
                
            question_text = str(row[0]).strip()
            question_type = str(row[1]).strip().lower()
            option_a = str(row[2]).strip() if row[2] else None
            option_b = str(row[3]).strip() if row[3] else None
            option_c = str(row[4]).strip() if row[4] else None
            option_d = str(row[5]).strip() if row[5] else None
            correct_answer = str(row[6]).strip()
            duration_seconds = int(row[7]) if row[7] else None
            group_name = str(row[8]).strip() if row[8] else None
            
            # Validate question type (only mcq and true_false allowed)
            if question_type not in ['mcq', 'true_false']:
                errors.append(f"Row {row_num}: Invalid question type '{question_type}'. Only 'mcq' and 'true_false' are allowed.")
                continue
            
            # Validate options based on question type
            if question_type == 'mcq':
                if not all([option_a, option_b, option_c, option_d]):
                    errors.append(f"Row {row_num}: MCQ questions require all 4 options")
                    continue
            
            elif question_type == 'true_false':
                if not option_a or not option_b:
                    errors.append(f"Row {row_num}: True/False questions require option A and option B")
                    continue
                # Clear options C and D for true/false
                option_c = None
                option_d = None
            
            # Get or validate group
            group = None
            if group_name:
                try:
                    group = QuestionGroup.objects.get(quiz=quiz, name=group_name)
                except QuestionGroup.DoesNotExist:
                    errors.append(f"Row {row_num}: Group '{group_name}' does not exist")
                    continue
            
            # Validate correct answer
            if correct_answer not in ['option_a', 'option_b', 'option_c', 'option_d']:
                errors.append(f"Row {row_num}: Correct answer must be option_a, option_b, option_c, or option_d")
                continue
            
            # Create question
            Question.objects.create(
                quiz=quiz,
                group=group,
                question_text=question_text,
                question_type=question_type,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_answer=correct_answer,
                duration_seconds=duration_seconds,
                order=row_num - 1
            )
            
            success_count += 1
            
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return success_count, errors
