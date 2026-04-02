import openpyxl, csv, io
from ..models import Question, QuestionGroup

VALID_TYPES = ('mcq', 'true_false')


def import_questions_from_file(file, quiz):
    """
    Import questions from Excel (.xlsx) or CSV file
    Returns dict with status, imported count, and errors
    """
    result = import_from_excel(file, quiz)
    
    if result['created'] > 0:
        return {
            'status': 'success',
            'imported': result['created'],
            'errors': result.get('errors', [])
        }
    else:
        return {
            'status': 'error',
            'message': '; '.join(result.get('errors', ['No questions imported'])),
            'imported': 0
        }


def import_from_excel(file, quiz):
    errors = []
    created = 0
    filename = getattr(file, 'name', '')

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
            rows = [r for r in reader]
            headers = reader.fieldnames or []
        else:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            raw_headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2):
                if any(c.value for c in row):
                    rows.append({raw_headers[i]: (str(c.value).strip() if c.value is not None else '') for i, c in enumerate(row)})
            headers = raw_headers
    except Exception as e:
        return {'created': 0, 'errors': [f'Cannot read file: {e}']}

    required = ['question_text', 'type', 'correct_answer']
    for col in required:
        if col not in [h.lower() for h in (headers or rows[0].keys() if rows else [])]:
            return {'created': 0, 'errors': [f'Missing required column: {col}']}

    next_order = quiz.questions.count() + 1

    for i, row in enumerate(rows, start=2):
        def g(k): return (row.get(k) or '').strip()

        text    = g('question_text')
        qtype   = g('type').lower()
        correct = g('correct_answer')

        if not text:
            errors.append(f'Row {i}: question_text empty'); continue
        if qtype not in VALID_TYPES:
            errors.append(f'Row {i}: invalid type "{qtype}"'); continue
        if not correct:
            errors.append(f'Row {i}: correct_answer empty'); continue

        dur = g('duration_seconds')
        duration = int(dur) if dur.isdigit() else None
        group_name = g('group_name')
        group = None
        if group_name:
            try:
                group = QuestionGroup.objects.get(quiz=quiz, name__iexact=group_name)
            except QuestionGroup.DoesNotExist:
                errors.append(f'Row {i}: group "{group_name}" not found'); continue

        Question.objects.create(
            quiz=quiz, group=group, question_text=text,
            question_type=qtype,
            option_a=g('option_a') or None, option_b=g('option_b') or None,
            option_c=g('option_c') or None, option_d=g('option_d') or None,
            correct_answer=correct, duration_seconds=duration, order=next_order)
        next_order += 1
        created += 1

    return {'created': created, 'errors': errors}
